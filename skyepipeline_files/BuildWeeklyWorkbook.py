#!/usr/bin/env python3
import pandas as pd
import numpy as np
import openpyxl.utils as get_column_letter
from openpyxl.styles import Font

"""
BuildWeeklyWorkbook.py

Purpose:
 - Create the final Excel workbook for the period report. Writes two tabs:
     1) `Master Log` (detailed per-order rows)
     2) `Financial Summary` (human-friendly, pre-formatted summary table)

Key operations performed:
 - Accepts `master` and `weekly_summary` as DataFrames or file paths.
 - Recalculates and validates key financial metrics to ensure consistency
     between the Master Log and the summary inputs.
 - Builds a readable summary table (rows with escaped leading `+`/`-` so
     Excel does not treat them as formulas) and autosizes columns for neat output.
 - Writes the two-sheet workbook to `output_path` using `openpyxl` engine.

Notes:
 - Cells that begin with `+` or `-` are escaped to avoid Excel formula parsing.
 - The module auto-sizes columns for readability and requires `openpyxl` to
     write `.xlsx` files.
"""

#TODO: CHANGE WEEKLY_ENDING_INVENTORY TO JUST ENDING INVENTORY THROUGHOUT

def escape_excel_formula(text):
    if isinstance(text, str) and text and text[0] in ("=", "+", "-"):
        return "'" + text
    return text

def autosize_columns(ws):
    for col in ws.columns:
        max_length = 0
        # First cell in the column
        first_cell = col[0]
        # Get the column letter directly from the cell
        col_letter = first_cell.column_letter

        for cell in col:
            try:
                cell_value = "" if cell.value is None else str(cell.value)
                if len(cell_value) > max_length:
                    max_length = len(cell_value)
            except Exception:
                pass

        # Add a little padding
        adjusted_width = max_length + 2
        ws.column_dimensions[col_letter].width = adjusted_width


def build_weekly_workbook(
    master,
    weekly_summary,
    output_path="Skye_Weekly_Report.xlsx",
    pos_bars=None,
    tot_pos_bars=None,
):
    """
    Build an Excel workbook from `master` and `weekly_summary` which may be
    file paths or DataFrames. Writes the Excel workbook to `output_path`.
    """
    # ---- LOAD DATA ----
    if isinstance(master, pd.DataFrame):
        master_df = master.copy()
    else:
        master_df = pd.read_csv(master)

    if isinstance(weekly_summary, pd.DataFrame):
        summary_raw = weekly_summary.copy()
    else:
        summary_raw = pd.read_csv(weekly_summary)

    if summary_raw.empty:
        raise ValueError("weekly_summary.csv is empty – run your summary script first.")

    s = summary_raw.iloc[0].copy()

    # ---- ENSURE MASTER NUMERIC COLUMNS ----
    for col in ["subtotal", "discount", "shipping", "tax", "bar_cogs", "total_shipping_cost", "total_bars_sold"]:
        if col in master_df.columns:
            master_df[col] = pd.to_numeric(master_df[col], errors="coerce")

    # ---- ENSURE SUMMARY NUMERIC COLUMNS ----
    num_cols = [
        "Gross_Revenue",  # may or may not be used
        "Taxes_Collected",
        "COGS_Total",
        "Shipping_Costs_Total",
        "Shipping_Costs_Orders",
        "Receiving_Sum",
        "Payment_Processing_Fee",
        "Gross_Profit",
        "Gross_Margin",
        "Starting_Inventory_Bars",
        "Boxes_Sold_This_Week",
        "Bars_Sold_This_Week",
        "Total_Inventory_Sold_Bars",
        "Weekly_Ending_Inventory_Bars",
    ]
    for col in num_cols:
        if col in summary_raw.columns:
            s[col] = pd.to_numeric(s[col], errors="coerce")

    # ======================================================
    #           RE-CALCULATE KEY FINANCIALS
    # ======================================================

    # Revenue (product only, net of discounts, EXCLUDING shipping)
    # Formula: sum(subtotal - discount)
    revenue_product = (summary_raw["Gross_Revenue"]).sum(skipna=True)

    # Shipping collected from customers (Shopify "shipping" column)
    shipping_collected = summary_raw["Shipping_Collected"].sum(skipna=True)

    # Gross Revenue = product revenue + shipping collected
    gross_revenue = revenue_product + shipping_collected

    # Taxes collected
    # Use master to stay consistent with actual orders
    taxes_collected = summary_raw["Taxes_Collected"].sum(skipna=True)

    # COGS total – use summary (already sum of bar_cogs)
    if "COGS_Total" in s.index and not pd.isna(s["COGS_Total"]):
        cogs_total = float(s["COGS_Total"])
    else:
        cogs_total = master_df["bar_cogs"].sum(skipna=True)

    # Total 3PL Costs = shipping (3PL) + receiving + payment processing fee
    # Already computed in weekly_summary as Shipping_Costs_Total
    if "Shipping_Costs_Total" in s.index and not pd.isna(s["Shipping_Costs_Total"]):
        total_3pl_costs = float(s["Shipping_Costs_Total"])
    else:
        total_3pl_costs = 0.0

    # Gross Profit = Gross Revenue - COGS - Total 3PL Costs
    gross_profit = gross_revenue - cogs_total - total_3pl_costs

    # Gross Margin = Gross Profit / Gross Revenue
    gross_margin = gross_profit / gross_revenue if gross_revenue != 0 else np.nan

    # ======================================================
    #              INVENTORY / UNIT NUMBERS
    # ======================================================

    starting_inventory = int(s.get("Starting_Inventory_Bars", 0) or 0)
    
    cases_sold = int(s.get("Cases_Sold_This_Week", 0) or 0)
    boxes_sold = int(s.get("Boxes_Sold_This_Week", 0) or 0)
    bars_sold = int(s.get("Bars_Sold_This_Week", 0) or 0)
    total_inventory_sold = int(s.get("Total_Inventory_Sold_Bars", 0) or 0)
    weekly_ending_inventory = int(s.get("Weekly_Ending_Inventory_Bars", 0) or 0)

    # ======================================================
    #           BUILD PRETTY SUMMARY TABLE
    # ======================================================

    rows = []

    # Header
    rows.append([
        escape_excel_formula("============== Cumulative Period Financials ====================="),
        ""
    ])

    # Revenue structure
    # Use numeric types for values so Excel stores real numbers and can be formatted
    rows.append([escape_excel_formula("Revenue"), revenue_product])
    rows.append([escape_excel_formula("+ Shipping collected"), shipping_collected])
    rows.append([escape_excel_formula("------------------------------------------------------"), ""])
    rows.append([escape_excel_formula("Gross Revenue"), gross_revenue])

    # Taxes
    rows.append([escape_excel_formula("+ Taxes Collected"), taxes_collected])

    # COGS & 3PL
    # Store subtractive amounts as negative numbers so Excel can format them
    try:
        cogs_value = -abs(float(cogs_total)) if not pd.isna(cogs_total) else np.nan
    except Exception:
        cogs_value = -abs(cogs_total) if cogs_total is not None else np.nan

    try:
        total_3pl_value = -abs(float(total_3pl_costs)) if not pd.isna(total_3pl_costs) else np.nan
    except Exception:
        total_3pl_value = -abs(total_3pl_costs) if total_3pl_costs is not None else np.nan

    rows.append([escape_excel_formula("- COGS"), cogs_value])
    rows.append([
        escape_excel_formula("- Total 3PL Costs (shipping, receiving, payment processing fee)"),
        total_3pl_value, "For for next period processing fee, lookup in pdf invoice"
    ])
    rows.append([escape_excel_formula("------------------------------------------------------"), ""])

    # Gross Profit & Margin
    rows.append([escape_excel_formula("Gross Profit"), gross_profit])
    if not np.isnan(gross_margin):
        # store as a fraction (e.g., 0.25 for 25%) so Excel percent formatting works
        rows.append([escape_excel_formula("Gross Margin"), gross_margin])
    else:
        rows.append([escape_excel_formula("Gross Margin"), "N/A"])

    # Spacer
    rows.append(["", ""])

    # Inventory header
    rows.append([escape_excel_formula("===============Inventory / Units======================="), ""])

    # Inventory details (store as integers so Excel keeps numeric types)

    rows.append([escape_excel_formula("Starting Inventory (bars)"), starting_inventory, ""])

    # Insert a row of '=' signs and a blank row before Boxes Sold
    rows.append([escape_excel_formula("======================================================"), "", ""])
    rows.append(["", "", ""])

    rows.append([escape_excel_formula("Cases Sold/Sent Out This Period"), cases_sold, ""])
    rows.append([escape_excel_formula("Boxes Sold/Sent Out This Period"), boxes_sold, ""])
    rows.append([escape_excel_formula("Single Bars Sold/Sent Out This Period"), bars_sold, ""])
    
    
    # separator (visual)
    rows.append(["","",""])
    rows.append([escape_excel_formula("------------------------------------------------------"), "", ""])


    # --- Concise double-check for cases, boxes, bars ---
    def get_mask(item):
        return (master_df.get("box_or_bar_or_case") == item)

    def not_sales_mask():
        email_col = master_df["email"].astype(str).str.strip().str.upper() if "email" in master_df.columns else ""
        source_col = master_df["source"].astype(str).str.strip().str.lower() if "source" in master_df.columns else ""
        sources_col = master_df["sources"].astype(str).str.strip().str.lower() if "sources" in master_df.columns else ""
        return (
            (email_col != "SENT TO SALES TEAM") & (source_col != "sales_team") & (sources_col != "sales_team")
        )

    # Bars per unit
    bars_per_case = 168
    bars_per_box = 7
    bars_per_single = 1

    # Master log counts
    mask = not_sales_mask()
    master_case_qty = int(pd.to_numeric(master_df.loc[get_mask("case") & mask, "line_item_quantity"], errors="coerce").sum(skipna=True) or 0)
    master_box_qty = int(pd.to_numeric(master_df.loc[get_mask("box") & mask, "line_item_quantity"], errors="coerce").sum(skipna=True) or 0)
    master_bar_qty = int(pd.to_numeric(master_df.loc[get_mask("bar") & mask, "line_item_quantity"], errors="coerce").sum(skipna=True) or 0)

    # Derived bars from master log
    master_case_bars = master_case_qty * bars_per_case
    master_box_bars = master_box_qty * bars_per_box
    master_bar_bars = master_bar_qty * bars_per_single

    # Derived bars from summary
    cases_sold_summary = int(s.get("Cases_Sold_This_Week", 0) or 0)
    boxes_sold_summary = int(s.get("Boxes_Sold_This_Week", 0) or 0)
    bars_sold_summary = int(s.get("Bars_Sold_This_Week", 0) or 0)
    summary_case_bars = cases_sold_summary * bars_per_case
    summary_box_bars = boxes_sold_summary * bars_per_box
    summary_bar_bars = bars_sold_summary * bars_per_single

    # Notes for mismatches
    note_case = f"Mismatch: master-derived bars={master_case_bars}" if master_case_bars != summary_case_bars else ""
    note_box = f"Mismatch: master-derived bars={master_box_bars}" if master_box_bars != summary_box_bars else ""
    note_bar = f"Mismatch: master-derived bars={master_bar_bars}" if master_bar_bars != summary_bar_bars else ""

    rows.append([escape_excel_formula("Case Bars Sold/Sent Out (case * 168 bars)"), summary_case_bars, note_case])
    rows.append([escape_excel_formula("+ Box Bars Sold/Sent Out (box * 7 bars)"), summary_box_bars, note_box])
    rows.append([escape_excel_formula("+ Single Bars Sold/Sent Out (single * 1 bar)"), summary_bar_bars, note_bar])

    # Add a row of dashes between single bars sold and total inventory sold
    rows.append([escape_excel_formula("------------------------------------------------------"), "", ""])

    rows.append([escape_excel_formula("Total Inventory Sold (bars)"), total_inventory_sold, ""])

    # Insert three blank rows, then a compact inventory summary block

    rows.append([escape_excel_formula("======================================================"), "", ""])
    rows.append(["", "", ""])  # blank

    rows.append([escape_excel_formula("Starting Inventory (bars)"), starting_inventory, ""])
    # Show total inventory sold as a subtractive value in the compact summary
    try:
        total_inventory_sold_value = -abs(int(total_inventory_sold))
    except Exception:
        total_inventory_sold_value = -abs(total_inventory_sold) if total_inventory_sold is not None else np.nan

    rows.append([escape_excel_formula("- Total Inventory Sold (bars)"), total_inventory_sold_value, ""])
    rows.append([escape_excel_formula("------------------------------------------------------"), "", ""])
    rows.append([escape_excel_formula("Ending Inventory (bars)"), weekly_ending_inventory, "Use this in next period Starting Inventory"])


    # Add a third column for optional notes/comments (e.g., mismatches)
    summary_pretty = pd.DataFrame(rows, columns=["Metric", "Value", "Note"])

    # ---- POS / 3PL Remaining Bars ----
    # Determine POS bars value: if caller provided `pos_bars` use it,
    # otherwise prompt the user when running interactively.
    try:
        if pos_bars is None:
            user_input = input("Enter Bars to be sold (POS) (integer, 0 if none): ")
            pos_bars_val = int(user_input.strip()) if str(user_input).strip() != "" else 0
        else:
            pos_bars_val = int(pos_bars)
    except Exception:
        pos_bars_val = 0

    # Add GTM / sales-team sendout bars into the POS bars count so they
    # are available to be subtracted by the POS calculation but still
    # remain tracked in the master log (use `exclude_from_bars_sold`).
    gtm_bars = 0
    if "exclude_from_bars_sold" in master_df.columns and "total_bars_sold" in master_df.columns:
        try:
            gtm_bars = int(master_df.loc[master_df["exclude_from_bars_sold"] == True, "total_bars_sold"].sum(skipna=True) or 0)
        except Exception:
            try:
                gtm_bars = int((master_df.loc[master_df["exclude_from_bars_sold"] == True, "total_bars_sold"].sum(skipna=True)))
            except Exception:
                gtm_bars = 0

    try:
        pos_bars_val = int(pos_bars_val) + int(gtm_bars)
    except Exception:
        pass

    
    
    # Bars sent total for POS = total sent to sales team + newly sent out gtm bars
    try:
        tot_for_pos = int(tot_pos_bars) + int(gtm_bars)
    except Exception:
        tot_pos_bars = 0
    
    # Bars left for POS = pos_bars - single bars sold (per request)  
    try:
        bars_left_for_pos = int(pos_bars_val) - int(bars_sold)
    except Exception:
        bars_left_for_pos = 0

    # Bars left at 3PL = ending inventory - bars_left_for_pos
    try:
        bars_left_at_3pl = int(weekly_ending_inventory) - int(tot_for_pos)
    except Exception:
        bars_left_at_3pl = 0

    pos_note = ""
    if bars_left_at_3pl < 0:
        pos_note = f"Negative at 3PL: {bars_left_at_3pl}"

    # Append POS rows to the DataFrame so they appear in the Financial Summary
    extra_rows = [
        [escape_excel_formula("======================================================"), "", ""],
        ["", "", ""],
        [escape_excel_formula("=============== POS / Remaining Inventory ============"), "", ""],
        [escape_excel_formula("Total POS Bars that were given to sales members"), tot_for_pos, "Use this for POS sent to sales members"],
        ["" ,"" ,""],
        [escape_excel_formula("Bars to be sold (POS)"), pos_bars_val, ""],
        [escape_excel_formula("- Single Bars Sold/Sent Out"), bars_sold, ""],
        [escape_excel_formula("------------------------------------------------------"), "", ""],
        [escape_excel_formula("Bars outstanding (POS)"), bars_left_for_pos, "Use this in next period POS bars"],
        [escape_excel_formula("======================================================"), "", ""],
        ["", "", ""],
        [escape_excel_formula("Ending Inventory (bars)"), weekly_ending_inventory],
        [escape_excel_formula("- Bars given out to Sales Members"), -abs(tot_for_pos), ""],
        [escape_excel_formula("------------------------------------------------------"), "", ""],
        [escape_excel_formula("Bars left at 3PL"), bars_left_at_3pl, pos_note],
    ]

    # concat extra rows onto summary_pretty
    if not extra_rows:
        pass
    else:
        extra_df = pd.DataFrame(extra_rows, columns=["Metric", "Value", "Note"])
        summary_pretty = pd.concat([summary_pretty, extra_df], ignore_index=True)

    # ======================================================
    #                WRITE EXCEL WITH TWO TABS
    # ======================================================

    with pd.ExcelWriter(output_path, engine="openpyxl") as writer:
        # Tab 1: master log table
        master_df.to_excel(writer, sheet_name="Master Log", index=False)
    
    # Tab 2: Financial & Inventory Summary
        summary_pretty.to_excel(writer, sheet_name="Financial Summary", index=False)

        wb = writer.book
        for sheet_name in wb.sheetnames:
            ws = wb[sheet_name]
            autosize_columns(ws)

        # Apply number formats to Financial Summary so values are stored as numbers
        if "Financial Summary" in wb.sheetnames:
            ws_summary = wb["Financial Summary"]
            # DataFrame wrote headers in row 1; data starts at row 2
            for row in ws_summary.iter_rows(min_row=2, min_col=1, max_col=3):
                metric_cell = row[0]
                value_cell = row[1]
                note_cell = row[2] if len(row) > 2 else None
                metric_text = str(metric_cell.value or "").lower()
                # Skip empty values and separators
                if value_cell.value is None or value_cell.value == "":
                    continue

                # Determine if this metric is subtractive (label starts with '-')
                is_subtractive = metric_text.strip().startswith("-")

                # Percentages (Gross Margin)
                if "margin" in metric_text:
                    try:
                        # Expecting a fraction (e.g., 0.25) -> display as percent
                        if is_subtractive:
                            value_cell.number_format = '0.00%;(0.00%)'
                        else:
                            value_cell.number_format = '0.00%'
                    except Exception:
                        pass
                # Currency-like values
                # avoid matching bare '3pl' so inventory labels like 'Bars left at 3PL'
                # are not treated as currency; match cost-related terms instead
                elif any(k in metric_text for k in ["revenue", "tax", "cogs", "shipping", "profit", "cost", "payment processing fee"]):
                    try:
                        # Use accounting-style formatting when subtractive so negatives show parentheses
                        if is_subtractive:
                            value_cell.number_format = '"$"#,##0.00;(\"$\"#,##0.00)'
                        else:
                            value_cell.number_format = '"$"#,##0.00'
                    except Exception:
                        pass
                # Integer counts (inventory / units)
                elif any(k in metric_text for k in ["inventory", "bars", "boxes", "sold", "starting inventory", "ending inventory", "total inventory"]):
                    try:
                        if is_subtractive:
                            value_cell.number_format = '#,##0;(#,##0)'
                        else:
                            value_cell.number_format = '#,##0'
                    except Exception:
                        pass

        print(f"Workbook written to: {output_path}")


# Runner for testing
# if __name__ == "__main__":
#     build_weekly_workbook(
#         master_log_path="master_log_Oct24_to_Nov21.csv",
#         weekly_summary_path="weekly_summary.csv",
#         output_path="Skye_Period_Report.xlsx",
#     )

